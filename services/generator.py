from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
from PIL import Image

from payroll.config import (
    TEMPLATE_FILE, TEMPLATE_SHEET, GAP_ROWS, CELL_MAP,
    BLOCK_TOP_LEFT, BLOCK_BOTTOM_RIGHT, LOGO_PATH
)
from payroll.data.loader import load_collaborators
from payroll.excel.copy_blocks import copy_range_between_sheets
from payroll.excel.fill_template import fill_block, paste_template_to_sheet_n_times
from payroll.excel.images import add_block_images

from payroll.utils.a1rc import a1_to_rc

def build_struct_sheets_and_fill(progress_cb, struct: dict,
                                 template_file: str,
                                 template_sheet: str,
                                 r1: int, c1: int, r2: int, c2: int,
                                 cell_map: dict,
                                 logo_path: str | None = None,
                                 logo_anchors=("A1","A15"),
                                 gap: int = 4,
                                 output_file: str = "collaborators_by_place.xlsx",
                                 tick_every: int = 10):
    tmpl_wb = load_workbook(template_file, data_only=False)
    if template_sheet not in tmpl_wb.sheetnames:
        raise ValueError(f"'{template_sheet}' no se ha encontrado {template_file}")
    ws_src = tmpl_wb[template_sheet]

    out_wb = Workbook()
    out_wb.remove(out_wb.active)

    height = r2 - r1 + 1
    total_receipts = 0
    for place, objs in struct.items():
        ws_dst = out_wb.create_sheet(title=(str(place)[:31] or "Sheet"))
        if not objs:
            progress_cb(f"‘{place}’: (sin registros)")
            continue

        N = len(objs)
        total_receipts += N
        progress_cb(f"‘{place}’: preparando {N} bloques…")

        # 1) Paste formatted template blocks (tick every few)
        for i in range(N):
            dest_r = 1 + i * (height + gap)
            copy_range_between_sheets(ws_src, r1, c1, r2, c2, ws_dst, dest_r, 1)
            if (i + 1) % tick_every == 0 or (i + 1) == N:
                progress_cb(f"‘{place}’: pegados {i+1}/{N} bloques")

        # 2) Fill each block (+ optional images), also ticking
        for i, obj in enumerate(objs):
            base_row = 1 + i * (height + gap)
            fill_block(ws_dst, base_row, 1, obj, cell_map)
            if "salary" in cell_map:
                _r, _c = a1_to_rc(cell_map["salary"])
                ws_dst.cell(row=base_row + _r - 1, column=_c).font = Font(color="FFFFFF")
            if logo_path:
                add_block_images(ws_dst, base_row, 1, logo_anchors, logo_path)
            if (i + 1) % tick_every == 0 or (i + 1) == N:
                progress_cb(f"‘{place}’: llenados {i+1}/{N} bloques")

        progress_cb(f"‘{place}’: listo ({N} duplicados).\n")

    out_wb.save(output_file)
    progress_cb(f"{total_receipts} recibos de nómina creados en {output_file}.")


def generate_payroll_receipts(data_file: str, progress_cb=print, out_filename: str | None = None, week: str | None = None, pay_date: str | None = None):
    struct = load_collaborators(data_file, week=week, pay_date=pay_date)

    r1, c1 = a1_to_rc(BLOCK_TOP_LEFT)
    r2, c2 = a1_to_rc(BLOCK_BOTTOM_RIGHT)

    build_struct_sheets_and_fill(
        progress_cb,
        struct=struct,
        template_file=TEMPLATE_FILE,
        template_sheet=TEMPLATE_SHEET,
        r1=r1, c1=c1, r2=r2, c2=c2,
        cell_map=CELL_MAP,
        logo_path=LOGO_PATH,
        logo_anchors=("A1", "A15"),
        gap=GAP_ROWS,
        output_file= (out_filename or f"{date.today()}.xlsx"),
    )
