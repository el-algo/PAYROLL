from copy import copy
from openpyxl.formula.translate import Translator

def copy_cell(src_cell, dst_cell) -> None:
    val = src_cell.value
    if isinstance(val, str) and val.startswith("="):
        try:
            val = Translator(val, origin=src_cell.coordinate).translate_formula(dst_cell.coordinate)
        except Exception:
            pass
    dst_cell.value = val

    if src_cell.has_style:
        dst_cell.font = copy(src_cell.font)
        dst_cell.fill = copy(src_cell.fill)
        dst_cell.border = copy(src_cell.border)
        dst_cell.alignment = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy(src_cell.protection)

    if src_cell.hyperlink:
        dst_cell._hyperlink = copy(src_cell._hyperlink)
    if src_cell.comment:
        dst_cell.comment = copy(src_cell.comment)

def src_merges_in_rect(ws_src, r1, c1, r2, c2):
    out = []
    for mr in ws_src.merged_cells.ranges:
        if (r1 <= mr.min_row <= r2 and r1 <= mr.max_row <= r2 and
            c1 <= mr.min_col <= c2 and c1 <= mr.max_col <= c2):
            out.append(mr)
    return out

def copy_dimensions(ws_src, ws_dst, r1, c1, r2, c2, row_off, col_off):
    from openpyxl.utils.cell import get_column_letter
    for c in range(c1, c2 + 1):
        src_letter = get_column_letter(c)
        dst_letter = get_column_letter(c + col_off)
        dim = ws_src.column_dimensions.get(src_letter)
        if dim and dim.width is not None:
            ws_dst.column_dimensions[dst_letter].width = dim.width
    for r in range(r1, r2 + 1):
        dim = ws_src.row_dimensions.get(r)
        if dim and dim.height is not None:
            ws_dst.row_dimensions[r + row_off].height = dim.height

def copy_range_between_sheets(ws_src, r1, c1, r2, c2, ws_dst, dest_r1, dest_c1):
    row_off = dest_r1 - r1
    col_off = dest_c1 - c1

    src_merges = src_merges_in_rect(ws_src, r1, c1, r2, c2)
    src_merge_map = {}
    for mr in src_merges:
        for rr in range(mr.min_row, mr.max_row + 1):
            for cc in range(mr.min_col, mr.max_col + 1):
                src_merge_map[(rr, cc)] = (mr.min_row, mr.min_col)

    for rr in range(r1, r2 + 1):
        for cc in range(c1, c2 + 1):
            if (rr, cc) in src_merge_map and src_merge_map[(rr, cc)] != (rr, cc):
                continue
            src = ws_src.cell(rr, cc)
            dst = ws_dst.cell(rr + row_off, cc + col_off)
            copy_cell(src, dst)

    copy_dimensions(ws_src, ws_dst, r1, c1, r2, c2, row_off, col_off)

    for mr in src_merges:
        ws_dst.merge_cells(
            start_row=mr.min_row + row_off, start_column=mr.min_col + col_off,
            end_row=mr.max_row + row_off, end_column=mr.max_col + col_off
        )
